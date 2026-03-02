import openpyxl
from openpyxl_image_loader import SheetImageLoader
import os
import re

wb = openpyxl.load_workbook("recursos_supcon/SUPCON Oversea Product List (1).xlsx", data_only=True)
os.makedirs("assets/supcon_img/mapped", exist_ok=True)

def sanitize_filename(name):
    if not name:
        return "unknown"
    name = str(name).strip()
    name = re.sub(r'[^A-Za-z0-9_ \-]', '', name)
    return name.replace(' ', '_').replace('__', '_')

for sheet_name in wb.sheetnames:
    print(f"--- Sheet: {sheet_name} ---")
    ws = wb[sheet_name]
    try:
        image_loader = SheetImageLoader(ws)
        for img in ws._images:
            if hasattr(img.anchor, '_from'):
                row = img.anchor._from.row
                col = img.anchor._from.col
            elif hasattr(img.anchor, 'from_'):
                row = img.anchor.from_.row
                col = img.anchor.from_.col
            else:
                continue

            excel_row = row + 1
            product_name = ws.cell(row=excel_row, column=2).value
            
            if not product_name:
                product_name = ws.cell(row=excel_row, column=1).value
            
            if not product_name:
                product_name = f"{sheet_name}_row_{excel_row}"
                
            safe_name = sanitize_filename(product_name)
            
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col + 1)
            cell_coord = f"{col_letter}{excel_row}"
            
            try:
                if image_loader.image_in(cell_coord):
                    pil_image = image_loader.get(cell_coord)
                    pil_image.save(f"assets/supcon_img/mapped/{safe_name}.png")
                    print(f"Saved: {safe_name}.png")
                else:
                    print(f"Image not found at {cell_coord} by loader.")
            except Exception as e:
                print(f"Error saving {safe_name}: {e}")
                
    except Exception as e:
        print(f"Error reading images in {sheet_name}: {e}")
